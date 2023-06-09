from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import utils


def create_columns_in_excel(column_titels: list, file_name: str) -> None:
    """Создает названия столцов в таблице Excel."""
    workbook = Workbook()
    rings = workbook.active
    for col_index, column_name in enumerate(column_titels, 1):
        cell = rings.cell(row=1, column=col_index, value=column_name)
        cell.font = Font(bold=True)
        rings.column_dimensions[cell.column_letter].bestFit = True

    workbook.save(file_name)


def formated_table(sheet: Worksheet, file_name: str) -> None:
    """Форматирование внешнего вида таблицы."""
    workbook = sheet.parent

    max_lengths = {}
    for row in sheet.iter_rows(values_only=True):
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value:
                cell_length = len(str(cell_value))
                max_lengths[col_idx] = max(
                    max_lengths.get(col_idx, 0), cell_length
                )

    for col_idx, max_length in max_lengths.items():
        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        if column_letter != 'E':
            adjusted_width = (max_length + 1) * 1.1
            sheet.column_dimensions[column_letter].width = adjusted_width
        if column_letter == 'E':
            sheet.column_dimensions[column_letter].width = max_length // 6.5

    for col_idx in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].bestFit = True

    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

    workbook.save(file_name)


def update_excel_data(data: dict, file_name: str) -> None:
    """Обновление информации в таблицы Ecxel."""
    print('Сохраняю информацию в Excel')

    workbook = load_workbook(file_name)

    rings = workbook.active
    columns_titles = [cell.value for cell in rings[1]]

    for num, ring in enumerate(data, start=2):
        ring_data_with_details = utils.make_advanced_details(ring)
        for key, value in ring_data_with_details.items():
            if key in columns_titles:

                column_index = columns_titles.index(key) + 1

                if isinstance(value, list):
                    value = '\n'.join(value)

                rings.cell(row=num, column=column_index, value=value)
            else:
                columns_titles.append(key)
                column_index = columns_titles.index(key) + 1

                column = rings.cell(row=1, column=column_index, value=key)
                column.font = Font(bold=True)

                if isinstance(value, list):
                    value = '\n'.join(value)

                rings.cell(row=num, column=column_index, value=value)

    formated_table(rings, file_name)
    workbook.save(file_name)
